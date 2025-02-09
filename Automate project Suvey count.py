import openpyxl
from datetime import datetime

# Load the Excel file
survey_file = openpyxl.load_workbook("Surveys Week 1 with false data.xlsx")
week_one = survey_file["Export"]

# Dictionaries to store survey counts per agent
total_surveys_per_agent = {}
total_satisfied_per_agent = {}
total_dissatisfied_per_agent = {}
total_neutral_per_agent = {}

# Dictionaries to store case numbers and comments per agent
satisfied_cases = {}
dissatisfied_cases = {}
neutral_cases = {}

# Map survey results to numerical values
survey_scores = {"Satisfied": 1, "Dissatisfied": -1, "Neutral": 0}

# Iterate through rows in the sheet
for row in range(2, week_one.max_row + 1):
    agent_name = week_one.cell(row, 1).value  # Agent Name
    case_number = week_one.cell(row, 2).value  # Case Number
    verbatim_comment = week_one.cell(row, 3).value  # Customer Comment
    survey_result = week_one.cell(row, 4).value  # Survey Result

    # Convert survey result into a numerical score
    if survey_result not in survey_scores:
        print(f"Warning: Survey result '{survey_result}' not recognized. Skipping this row.")
        continue  # Skip unrecognized survey result
    survey_result = survey_scores[survey_result]

    # Count total surveys per agent
    total_surveys_per_agent[agent_name] = total_surveys_per_agent.get(agent_name, 0) + 1

    # Store case numbers and comments based on survey type
    if survey_result == 1:  # Satisfied
        total_satisfied_per_agent[agent_name] = total_satisfied_per_agent.get(agent_name, 0) + 1
        satisfied_cases.setdefault(agent_name, []).append((case_number, verbatim_comment))
    elif survey_result == -1:  # Dissatisfied
        total_dissatisfied_per_agent[agent_name] = total_dissatisfied_per_agent.get(agent_name, 0) + 1
        dissatisfied_cases.setdefault(agent_name, []).append((case_number, verbatim_comment))
    else:  # Neutral
        total_neutral_per_agent[agent_name] = total_neutral_per_agent.get(agent_name, 0) + 1
        neutral_cases.setdefault(agent_name, []).append((case_number, verbatim_comment))

# Generate a timestamped filename
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
file_name = f"Survey_Results_Week_1_Separated_{timestamp}.xlsx"

# Create a new Excel workbook
new_workbook = openpyxl.Workbook()

# Create sheets for each section
total_surveys_sheet = new_workbook.active
total_surveys_sheet.title = "Total Surveys per Agent"
total_surveys_sheet.append(["Agent Name", "Total Surveys"])

# Write total surveys data, sorted alphabetically by agent name
for agent in sorted(total_surveys_per_agent.keys()):
    total_surveys_sheet.append([agent, total_surveys_per_agent[agent]])

# Add a sheet for satisfied surveys
satisfied_sheet = new_workbook.create_sheet("Satisfied Surveys")
satisfied_sheet.append(["Agent Name", "Satisfied Surveys", "Case Number", "Comment"])

# Write satisfied survey data, sorted alphabetically by agent name
for agent in sorted(total_satisfied_per_agent.keys()):
    satisfied_sheet.append([agent, total_satisfied_per_agent[agent]])  # Write total satisfied surveys
    for case, comment in satisfied_cases.get(agent, []):
        satisfied_sheet.append([agent, total_satisfied_per_agent[agent], case, comment])  # Write case number and comment

# Add a sheet for dissatisfied surveys
dissatisfied_sheet = new_workbook.create_sheet("Dissatisfied Surveys")
dissatisfied_sheet.append(["Agent Name", "Dissatisfied Surveys", "Case Number", "Comment"])

# Write dissatisfied survey data, sorted alphabetically by agent name
for agent in sorted(total_dissatisfied_per_agent.keys()):
    dissatisfied_sheet.append([agent, total_dissatisfied_per_agent[agent]])  # Write total dissatisfied surveys
    for case, comment in dissatisfied_cases.get(agent, []):
        dissatisfied_sheet.append([agent, total_dissatisfied_per_agent[agent], case, comment])  # Write case number and comment

# Add a sheet for neutral surveys
neutral_sheet = new_workbook.create_sheet("Neutral Surveys")
neutral_sheet.append(["Agent Name", "Neutral Surveys", "Case Number", "Comment"])

# Write neutral survey data, sorted alphabetically by agent name
for agent in sorted(total_neutral_per_agent.keys()):
    neutral_sheet.append([agent, total_neutral_per_agent[agent]])  # Write total neutral surveys
    for case, comment in neutral_cases.get(agent, []):
        neutral_sheet.append([agent, total_neutral_per_agent[agent], case, comment])  # Write case number and comment

# Save the new workbook to a file
new_workbook.save(file_name)

print(f"New Excel file '{file_name}' created successfully!")
